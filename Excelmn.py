import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os

wb=load_workbook('C:\\Users\\jegonzalez\\Documents\PYTHON\\Excelmanag\\managed.xlsx')
print(wb.worksheets)

ws=wb['Sheet1']

#list=[(ws.cell(1,1).value),(ws.cell(2,1).value)]
list1=[]


for i in range(1,6):
    list1.append(ws.cell(i,1).value)


#list.append(ws.cell(3,1).value)
#ws.cell(4,2).value="Como te encuentras"
#print(list)

result=[]
for i in list1:
      result.append(i)

print(result)
#valores=''.join(map(str,result))


valores2=''.join(map(lambda x: str(x)+ '|',result))
print(valores2)


#wb.save("C:\\Users\\jegonzalez\\Documents\PYTHON\\Excelmanag\\sample.xlsx")
