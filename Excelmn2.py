import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os

wb=load_workbook('C:\\Users\\jegonzalez\\Documents\PYTHON\\Excelmanag\\managed.xlsx')
print(wb.worksheets)

ws=wb['Sheet1']
#list=[(ws.cell(1,1).value),(ws.cell(2,1).value)]

list1=[]

i=1
j=2
#for value in ws.range('A'):
while (ws.cell(i,1).value != "OK"):
    if (ws.cell(i,1).value == ws.cell(j,1).value):
        ws.cell(i,3).value=ws.cell(i,2).value + '|'+ ws.cell(j,2).value
        #list1.append(ws.cell(i, 2).value)
        #list1.append(ws.cell(j, 2).value)
    i = i + 1
    j = j + 1



# for i in range(1,12):
#     for j in range(2,12):
#         if (ws.cell(i,1).value == ws.cell(j,1).value):
#             list1.append(ws.cell(i,2).value)
#             list1.append(ws.cell(j, 2).value)
#
# print(list1)

#result=[]
#for i in list1:
#      result.append(i)

#print(result)

#valores2=''.join(map(lambda x: str(x)+ '|',result))

#print(valores2)




#list.append(ws.cell(3,1).value)
#ws.cell(4,2).value="Como te encuentras"
#print(list)

#valores=''.join(map(str,result))

#ws.cell(3,3).value=valores2

wb.save("C:\\Users\\jegonzalez\\Documents\PYTHON\\Excelmanag\\sample.xlsx")





